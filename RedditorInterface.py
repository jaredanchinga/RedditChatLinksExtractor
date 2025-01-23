from datetime import datetime, timezone
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import DateEntry
import json
import os
import time
import praw
from openpyxl import load_workbook, Workbook
import sys
from PIL import Image, ImageTk

class RedditorInterface:
    def __init__(self, root):
        self.root = root
        self.root.title("Reddit Chat Links Extractor")
        
        # Create main container with two columns
        self.main_frame = ttk.Frame(root, padding="5")
        self.main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Create left frame for controls
        self.left_frame = ttk.Frame(self.main_frame)
        self.left_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=5)

        # Create right frame for terminal
        self.right_frame = ttk.Frame(self.main_frame)
        self.right_frame.grid(row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S), padx=5)

        # Working Directory Frame
        self.dir_frame = ttk.LabelFrame(self.left_frame, text="Working Directory", padding="2")
        self.dir_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=2)
        
        self.dir_var = tk.StringVar(value=os.getcwd())
        self.dir_entry = ttk.Entry(self.dir_frame, textvariable=self.dir_var, width=40)
        self.dir_entry.grid(row=0, column=0, padx=5, sticky=tk.W+tk.E)
        
        ttk.Button(self.dir_frame, text="Browse", command=self.browse_directory).grid(
            row=0, column=1, padx=5)
        
        # Date Range Frame
        self.date_frame = ttk.LabelFrame(self.left_frame, text="Date Range", padding="2")
        self.date_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=2)
        
        ttk.Label(self.date_frame, text="Start Date:").grid(row=0, column=0, padx=5)
        self.start_date = DateEntry(self.date_frame, width=12, background='darkblue',
                                  foreground='white', borderwidth=2)
        self.start_date.grid(row=0, column=1, padx=5)
        
        ttk.Label(self.date_frame, text="End Date:").grid(row=0, column=2, padx=5)
        self.end_date = DateEntry(self.date_frame, width=12, background='darkblue',
                                foreground='white', borderwidth=2)
        self.end_date.grid(row=0, column=3, padx=5)
        
        # Search Limits Frame
        self.limits_frame = ttk.LabelFrame(self.left_frame, text="Search Limits", padding="2")
        self.limits_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=2)
        
        ttk.Label(self.limits_frame, text="Max Posts/Subreddit:").grid(row=0, column=0, padx=5)
        self.max_posts = ttk.Entry(self.limits_frame, width=8)
        self.max_posts.insert(0, "500")
        self.max_posts.grid(row=0, column=1, padx=5)
        
        ttk.Label(self.limits_frame, text="Max Comments/Post:").grid(row=0, column=2, padx=5)
        self.max_comments = ttk.Entry(self.limits_frame, width=8)
        self.max_comments.insert(0, "500")
        self.max_comments.grid(row=0, column=3, padx=5)
        
        ttk.Label(self.limits_frame, text="Max Comment Depth:").grid(row=0, column=4, padx=5)
        self.max_depth = ttk.Entry(self.limits_frame, width=8)
        self.max_depth.insert(0, "100")
        self.max_depth.grid(row=0, column=5, padx=5)
        
        # Region Selection Frame with "Select All" option
        self.region_frame = ttk.LabelFrame(self.left_frame, text="Region Selection", padding="2")
        self.region_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=2)
        
        self.regions = ["ASIA", "CANADA", "EUROPE", "OCEANIA", "USA"]
        self.region_vars = {}
        
        # Add Select All checkbox for regions
        self.select_all_regions_var = tk.BooleanVar()
        ttk.Checkbutton(self.region_frame, text="Select All Regions", 
                       variable=self.select_all_regions_var,
                       command=self.toggle_all_regions).grid(
            row=0, column=0, columnspan=3, padx=5, sticky=tk.W)
        
        for i, region in enumerate(self.regions):
            var = tk.BooleanVar()
            ttk.Checkbutton(self.region_frame, text=region, variable=var,
                          command=self.check_region_selection).grid(
                row=(i//3)+1, column=i%3, padx=5, sticky=tk.W)
            self.region_vars[region] = var
        
        # Link Formats Frame with "Select All" option
        self.formats_frame = ttk.LabelFrame(self.left_frame, text="Link Formats", padding="2")
        self.formats_frame.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=2)
        
        self.link_formats = ['discord.gg', 'chat.whatsapp.com', 't.me', 'linktr.ee', 
                           'docs.google.com', 'groupme.com', 't.snapchat.com', 'ig.me', 'm.me']
        self.format_vars = {}
        
        # Add Select All checkbox for formats
        self.select_all_formats_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(self.formats_frame, text="Select All Formats", 
                       variable=self.select_all_formats_var,
                       command=self.toggle_all_formats).grid(
            row=0, column=0, columnspan=3, padx=5, sticky=tk.W)
        
        for i, format in enumerate(self.link_formats):
            var = tk.BooleanVar(value=True)
            ttk.Checkbutton(self.formats_frame, text=format, variable=var,
                          command=self.check_format_selection).grid(
                row=(i//3)+1, column=i%3, padx=5, sticky=tk.W)
            self.format_vars[format] = var

        # Control Buttons
        self.button_frame = ttk.Frame(self.left_frame)
        self.button_frame.grid(row=5, column=0, columnspan=2, pady=5)
        
        ttk.Button(self.button_frame, text="Save Settings", command=self.save_settings).grid(
            row=0, column=0, padx=5)
        ttk.Button(self.button_frame, text="Start Search", command=self.start_search).grid(
            row=0, column=1, padx=5)

        # Add logo at the bottom left
        try:
            if getattr(sys, 'frozen', False):
                script_dir = os.path.dirname(sys.executable)
            else:
                script_dir = os.path.dirname(os.path.abspath(__file__))
                
            logo_path = os.path.join(script_dir, 'logo.png')
            
            # Use PIL to load and resize the image
            pil_image = Image.open(logo_path)
            # Calculate size for icon (e.g., 64x64)
            icon_size = (64, 64)
            pil_image.thumbnail(icon_size)
            
            # Convert PIL image for Tkinter
            self.logo_image = ImageTk.PhotoImage(pil_image)
            self.root.iconphoto(True, self.logo_image)
            
            # Create another copy for the label
            label_size = (80, 80)  # Slightly smaller for bottom placement
            pil_image_label = Image.open(logo_path)
            pil_image_label.thumbnail(label_size)
            self.logo_label_image = ImageTk.PhotoImage(pil_image_label)
            
            # Add logo to interface at bottom left
            self.logo_label = ttk.Label(self.left_frame, image=self.logo_label_image)
            self.logo_label.grid(row=6, column=0, pady=(10,5), sticky=tk.W)  # Row 6 is after buttons
            
        except Exception as e:
            print(f"Warning: Could not load logo: {str(e)}")

        # Add terminal-like display on the right
        self.terminal = tk.Text(self.right_frame, height=30, width=60, 
                              bg='black', fg='white',
                              font=('Courier', 9))
        self.terminal.pack(fill=tk.BOTH, expand=True)
        
        # Add scrollbar to terminal
        self.terminal_scroll = ttk.Scrollbar(self.right_frame, 
                                           command=self.terminal.yview)
        self.terminal_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.terminal.config(yscrollcommand=self.terminal_scroll.set)
        
        # Modify print to write to both console and terminal
        original_print = print
        def print_to_terminal(*args, **kwargs):
            message = ' '.join(str(arg) for arg in args)
            self.terminal.insert(tk.END, message + '\n')
            self.terminal.see(tk.END)
            self.terminal.update()
            original_print(*args, **kwargs)  # Keep console output too
        
        import builtins
        builtins.print = print_to_terminal

        # Load saved settings if they exist
        self.load_settings()

    def toggle_all_regions(self):
        """Toggle all region checkboxes"""
        state = self.select_all_regions_var.get()
        for var in self.region_vars.values():
            var.set(state)

    def toggle_all_formats(self):
        """Toggle all format checkboxes"""
        state = self.select_all_formats_var.get()
        for var in self.format_vars.values():
            var.set(state)

    def check_region_selection(self):
        """Update 'Select All' checkbox based on individual selections"""
        all_selected = all(var.get() for var in self.region_vars.values())
        self.select_all_regions_var.set(all_selected)

    def check_format_selection(self):
        """Update 'Select All' checkbox based on individual selections"""
        all_selected = all(var.get() for var in self.format_vars.values())
        self.select_all_formats_var.set(all_selected)

    def save_settings(self):
        settings = {
            'working_directory': self.dir_var.get(),
            'start_date': self.start_date.get_date().strftime('%Y-%m-%d'),
            'end_date': self.end_date.get_date().strftime('%Y-%m-%d'),
            'max_posts': self.max_posts.get(),
            'max_comments': self.max_comments.get(),
            'max_depth': self.max_depth.get(),
            'regions': {region: var.get() for region, var in self.region_vars.items()},
            'formats': {format: var.get() for format, var in self.format_vars.items()}
        }
        
        settings_path = os.path.join(os.getcwd(), 'redditor_settings.json')
        with open(settings_path, 'w') as f:
            json.dump(settings, f)
        
        print(f"Settings saved to: {settings_path}")
        messagebox.showinfo("Success", "Settings saved successfully!")

    def load_settings(self):
        try:
            settings_path = os.path.join(os.getcwd(), 'redditor_settings.json')
            with open(settings_path, 'r') as f:
                settings = json.load(f)
            
            if 'working_directory' in settings:
                self.dir_var.set(settings['working_directory'])
                if os.path.exists(settings['working_directory']):
                    os.chdir(settings['working_directory'])
                    print(f"Working directory loaded: {settings['working_directory']}")
            
            self.start_date.set_date(datetime.strptime(settings['start_date'], '%Y-%m-%d'))
            self.end_date.set_date(datetime.strptime(settings['end_date'], '%Y-%m-%d'))
            self.max_posts.delete(0, tk.END)
            self.max_posts.insert(0, settings['max_posts'])
            self.max_comments.delete(0, tk.END)
            self.max_comments.insert(0, settings['max_comments'])
            self.max_depth.delete(0, tk.END)
            self.max_depth.insert(0, settings['max_depth'])
            
            for region, value in settings['regions'].items():
                if region in self.region_vars:
                    self.region_vars[region].set(value)
            
            for format, value in settings['formats'].items():
                if format in self.format_vars:
                    self.format_vars[format].set(value)
            
            print("Settings loaded successfully")
                    
        except FileNotFoundError:
            print("No settings file found. Using default values.")
            pass  # Use default values if no settings file exists

    def start_search(self):
        try:
            from Redditor import (read_subreddits, search_subreddit, 
                                update_excel_sheet, generate_report, 
                                remove_duplicates_keep_first, verify_credentials,
                                REDDIT_CREDENTIALS_LIST, load_proxies, get_next_proxy,
                                SEARCH_KEYWORDS)
            
            # Load proxies
            proxy_list = load_proxies()
            if not proxy_list:
                print("No proxies available. Please add proxies to proxies.txt")
                return
            
            # Verify credentials
            print("Verifying Reddit credentials...")
            valid_credentials = []
            for i, creds in enumerate(REDDIT_CREDENTIALS_LIST, 1):
                try:
                    reddit = praw.Reddit(**creds)
                    is_valid, error = verify_credentials(reddit)
                    if is_valid:
                        valid_credentials.append(creds)
                        print(f"✓ Credential set {i} ({creds['username']}) verified successfully")
                    else:
                        print(f"✗ Credential set {i} ({creds['username']}) failed: {error}")
                except Exception as e:
                    print(f"✗ Credential set {i} ({creds['username']}) failed: {str(e)}")
            
            if not valid_credentials:
                print("No valid credentials found. Please check your Reddit credentials.")
                return
            
            # Read subreddits
            all_subreddits = read_subreddits()
            selected_regions = [region for region, var in self.region_vars.items() if var.get()]
            
            if not selected_regions:
                messagebox.showerror("Error", "Please select at least one region")
                return
            
            # Filter subreddits based on selected regions
            subreddits = {region: all_subreddits[region] for region in selected_regions}
            
            # Initialize variables
            link_dict = {region: {format: set() for format in self.link_formats 
                                if self.format_vars[format].get()} 
                        for region in selected_regions}
            skipped_subreddits = []
            
            start_time = time.time()
            total_subreddits = sum(len(subs) for subs in subreddits.values())
            processed_count = 0
            current_credential_index = 0
            requests_made = 0
            current_proxy = None

            for region, region_subreddits in subreddits.items():
                print(f"\nProcessing region: {region}")
                excel_path = os.path.join(os.getcwd(), f'links{region}.xlsx')
                
                try:
                    book = load_workbook(excel_path)
                    print(f"Loaded existing file: {excel_path}")
                except FileNotFoundError:
                    book = Workbook()
                    book.remove(book.active)
                    print(f"Created new workbook for: {excel_path}")

                # Create sheets for each format if they don't exist
                for format in self.link_formats:
                    if format not in book.sheetnames:
                        book.create_sheet(format)
                        book[format].append(['Link', 'Subreddit', 'Date Added'])

                for subreddit in region_subreddits:
                    # Check if we need to rotate credentials and get new proxy
                    if requests_made >= 55:  # REQUESTS_PER_CREDENTIAL
                        current_credential_index = (current_credential_index + 1) % len(valid_credentials)
                        requests_made = 0
                        current_proxy = get_next_proxy(proxy_list)
                        if not current_proxy:
                            print("No more proxies available! Please add more proxies.")
                            return
                        print(f"\nRotating to credential set {current_credential_index + 1}")
                        time.sleep(2)
                    
                    # Initialize Reddit instance with current credentials and proxy
                    try:
                        if current_proxy is None:
                            current_proxy = get_next_proxy(proxy_list)
                            if not current_proxy:
                                print("No more proxies available!")
                                return

                        current_creds = valid_credentials[current_credential_index].copy()
                        current_creds['proxy'] = current_proxy
                        reddit = praw.Reddit(**current_creds)
                    except Exception as e:
                        print(f"Error initializing Reddit instance: {str(e)}")
                        continue
                    
                    processed_count += 1
                    progress = (processed_count / total_subreddits) * 100
                    print(f"Searching {subreddit} ({processed_count}/{total_subreddits})")
                    
                    links, error = search_subreddit(
                        reddit, 
                        subreddit,
                        SEARCH_KEYWORDS,
                        datetime.combine(
                            self.start_date.get_date(), 
                            datetime.min.time(),
                            tzinfo=timezone.utc
                        ),
                        datetime.combine(
                            self.end_date.get_date(), 
                            datetime.max.time(),
                            tzinfo=timezone.utc
                        )
                    )
                    requests_made += 1
                    
                    if error:
                        print(f"Skipped {subreddit}. Reason: {error}")
                        skipped_subreddits.append((subreddit, error))
                    else:
                        print(f"Found {len(links)} links in {subreddit}")
                        for link in links:
                            for format in self.link_formats:
                                if format in link and self.format_vars[format].get():
                                    update_excel_sheet(book[format], [link], subreddit)
                                    link_dict[region][format].add(link)
                                    break
                    
                    time.sleep(2)  # Base delay between requests

                book.save(excel_path)
                print(f"Saved all links for {region} to the Excel file: {excel_path}")
                remove_duplicates_keep_first(excel_path)

            end_time = time.time()
            report = generate_report(link_dict, skipped_subreddits, start_time, end_time, 
                                   self.start_date.get_date(), self.end_date.get_date())
            
            print("\nSearch completed!")
            print(report)
            messagebox.showinfo("Complete", "Search process completed successfully!")

        except Exception as e:
            print(f"Error: {str(e)}")
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    def create_default_files(self, directory):
        """Create default file structure in selected directory"""
        # Create Files directory if it doesn't exist
        files_dir = os.path.join(directory, 'Files')
        os.makedirs(files_dir, exist_ok=True)
        
        # Create default SUBREDDITS.xlsx if it doesn't exist
        subreddits_path = os.path.join(files_dir, 'SUBREDDITS.xlsx')
        if not os.path.exists(subreddits_path):
            wb = Workbook()
            # Create sheets for each region
            regions = ['ASIA', 'CANADA', 'EUROPE', 'OCEANIA', 'USA']
            for region in regions:
                if region not in wb.sheetnames:
                    wb.create_sheet(region)
            if 'Sheet' in wb.sheetnames:  # Remove default sheet
                wb.remove(wb['Sheet'])
            wb.save(subreddits_path)
            print(f"Created template SUBREDDITS.xlsx in {files_dir}")
        
        # Create default proxies.txt if it doesn't exist
        proxies_path = os.path.join(files_dir, 'proxies.txt')
        if not os.path.exists(proxies_path):
            with open(proxies_path, 'w') as f:
                f.write("# Add your proxies here, one per line\n")
                f.write("# Format: ip:port or username:password@ip:port\n")
            print(f"Created template proxies.txt in {files_dir}")

    def browse_directory(self):
        """Browse for working directory"""
        dir_path = filedialog.askdirectory(
            initialdir=self.dir_var.get(),
            title="Select Working Directory"
        )
        if dir_path:  # If a directory was selected
            self.dir_var.set(dir_path)
            os.chdir(dir_path)  # Change working directory
            print(f"Working directory changed to: {dir_path}")
            
            # Create default files if they don't exist
            self.create_default_files(dir_path)
            
            # Immediately try to load settings from the new directory
            self.load_settings()

if __name__ == "__main__":
    root = tk.Tk()
    app = RedditorInterface(root)
    root.mainloop() 