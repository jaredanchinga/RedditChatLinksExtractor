import re
import time
import os
from datetime import datetime, timedelta, timezone
import praw
import prawcore
from openpyxl import load_workbook, Workbook
import pandas as pd

# Constants
SCRIPT_DIR = os.getcwd()  # This will use the directory set by the interface
SUBREDDITS_FILE = os.path.join(SCRIPT_DIR, 'SUBREDDITS.xlsx')
LINK_FORMATS = ['discord.gg', 'chat.whatsapp.com', 't.me', 'linktr.ee', 'docs.google.com', 'groupme.com', 't.snapchat.com', 'ig.me', 'm.me']
SEARCH_KEYWORDS = ['groupchat', 'chat', 'discord', 'whatsapp', 'Telegram', 'snapchat', 'linktree', 'docs.google.com', 'groupme', 'discord.gg', 'chat.whatsapp.com', 't.me', 'snapchat.com', 'linktr.ee', 'Instagram Messenger', 'ig.me', 'm.me']

# Date range for search (easily editable) 
START_DATE = datetime(2025, 1, 1, tzinfo=timezone.utc)  # YYYY-MM-DD
END_DATE = datetime(2025, 12, 31, tzinfo=timezone.utc)   # YYYY-MM-DD

# Constants for rate limiting
REQUESTS_PER_CREDENTIAL = 55  # Maximum requests before switching credentials
MAX_POSTS_PER_SUBREDDIT = 500  # Maximum posts to fetch per subreddit
MAX_COMMENTS_PER_POST = 500    # Maximum comments to process per post
MAX_COMMENT_DEPTH = 100        # Maximum depth for nested comments

# Store credentials in a list
REDDIT_CREDENTIALS_LIST = [
    {
        'client_id': 'nJtJY5BQ6_wcmKyCEmD_FQ',
        'client_secret': '3hWgH6KR7eIsoKlkMBzOEE4YAUKglw',
        'user_agent': 'GMC 1',
        'redirect_uri': 'http://localhost:8080',
        'username': 'guidemyclass1',
        'password': '#Batch254',
    },
    {
        'client_id': 'vl4YNFvG5qnbxHbObXo6UQ',
        'client_secret': 'g2WuC-HjgL4wP5CMjpS1wVhweT-mFA',
        'user_agent': 'GMC 2',
        'redirect_uri': 'http://localhost:8080',
        'username': 'FinanceLarge8634',
        'password': '#Batch254',
    },
    {
        'client_id': 'OUHekWJ3pu0QteCsOZqXgg',
        'client_secret': 'bUt_X8gfg2msYOeyVRW-oxt28VxjUw',
        'user_agent': 'GMC 3',
        'redirect_uri': 'http://localhost:8080',
        'username': 'guidemyclass2',
        'password': '#Batch254',
    },
    {
        'client_id': 'gL-GrO7hGqg0VHqOfdfRtg',
        'client_secret': 'stIC7Bk7NB8RbR_mfvbCACr-volB2A',
        'user_agent': 'GMC 4',
        'redirect_uri': 'http://localhost:8080',
        'username': 'guidemyclass3',
        'password': '#Batch254',
    },
    {
        'client_id': '7d9hVCGPwJ5i_hNvCfvbvw',
        'client_secret': 'tpob_ivFCFK0iZY9z6iBvrYwUN44ow',
        'user_agent': 'GMC 5',
        'redirect_uri': 'http://localhost:8080',
        'username': 'guidemyclass4',
        'password': '#Batch254',
    }
]

log_callback = None

def set_log_callback(callback):
    global log_callback
    log_callback = callback

def read_subreddits():
    subreddits = {}
    excel_path = os.path.join(SCRIPT_DIR, 'SUBREDDITS.xlsx')
    try:
        wb = load_workbook(excel_path)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            subreddits[sheet_name] = [cell.value for cell in sheet['A'] if cell.value]
        message = f"Successfully read subreddits from {excel_path}"
        print(message)
        if log_callback:
            log_callback(message)
        for region, subs in subreddits.items():
            message = f"  {region}: {len(subs)} subreddits"
            print(message)
            if log_callback:
                log_callback(message)
        return subreddits
    except FileNotFoundError:
        message = f"Error: The file {excel_path} was not found."
        print(message)
        if log_callback:
            log_callback(message)
        message = f"Working directory: {SCRIPT_DIR}"
        print(message)
        if log_callback:
            log_callback(message)
        raise
    except Exception as e:
        print(f"An error occurred while reading the SUBREDDITS file: {str(e)}")
        raise

def extract_links(text):
    url_pattern = re.compile(r"http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\\(\\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+")
    matches = re.findall(url_pattern, text)
    formatted_links = []
    for link in matches:
        if link.count('https://') > 1:
            link = 'https://' + link.split('https://')[-1]
        formatted_links.append(link)
    return formatted_links

def search_subreddit(reddit, subreddit_name, keywords, start_date, end_date):
    try:
        subreddit = reddit.subreddit(subreddit_name)
        search_results = []
        
        print(f"Searching subreddit: {subreddit_name}")
        post_count = 0
        for post in subreddit.search(' OR '.join(keywords), limit=MAX_POSTS_PER_SUBREDDIT, sort='new'):
            post_date = datetime.fromtimestamp(post.created_utc, tz=timezone.utc)
            if start_date <= post_date <= end_date:
                links = extract_links(post.title + ' ' + post.selftext)
                search_results.extend(links)
                
                post.comments.replace_more(limit=None)
                
                def process_comments(comments, depth=0):
                    if depth >= MAX_COMMENT_DEPTH:
                        return []
                    comment_links = []
                    comment_count = 0
                    for comment in comments:
                        if comment_count >= MAX_COMMENTS_PER_POST:
                            break
                        comment_date = datetime.fromtimestamp(comment.created_utc, tz=timezone.utc)
                        if start_date <= comment_date <= end_date:
                            comment_links.extend(extract_links(comment.body))
                            comment_links.extend(process_comments(comment.replies, depth + 1))
                        comment_count += 1
                    return comment_links
                
                comment_links = process_comments(post.comments)
                search_results.extend(comment_links)
                post_count += 1
            elif post_date < start_date:
                break
        
        print(f"Processed {post_count} posts in {subreddit_name}")
        print(f"Total links found in {subreddit_name}: {len(search_results)}")
        return search_results, None
    except prawcore.exceptions.Forbidden:
        return [], "Access is forbidden. This may be a private subreddit."
    except prawcore.exceptions.NotFound:
        return [], "Subreddit not found. It may have been banned or may not exist."
    except praw.exceptions.RedditAPIException as e:
        return [], f"Reddit API error: {str(e)}"
    except Exception as e:
        return [], f"An unexpected error occurred: {str(e)}"

def update_excel_sheet(sheet, new_links, subreddit):
    existing_links = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        existing_links.add(row[0])

    new_links_added = 0
    current_date = datetime.now().strftime('%Y-%m-%d')
    for link in new_links:
        if link not in existing_links:
            sheet.append([link, subreddit, current_date])
            existing_links.add(link)
            new_links_added += 1

    print(f"Added {new_links_added} new links to sheet {sheet.title}")
    return new_links_added

def generate_report(link_dict, skipped_subreddits, start_time, end_time, start_date, end_date):
    report = f"Reddit Group Chat Scraper Report\n"
    report += f"Generated on: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}\n"
    report += f"Execution time: {end_time - start_time:.2f} seconds\n"
    report += f"Date range searched: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}\n\n"
    
    total_links = sum(len(links) for region_links in link_dict.values() for links in region_links.values())
    report += f"Total links found: {total_links}\n\n"
    
    for region, formats in link_dict.items():
        report += f"{region}:\n"
        for format, links in formats.items():
            report += f"  {format}: {len(links)} links\n"
        report += "\n"

    report += f"\nSkipped Subreddits:\n"
    for subreddit, reason in skipped_subreddits:
        report += f"  {subreddit}: {reason}\n"
    
    timestamp = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
    report_filename = os.path.join(SCRIPT_DIR, f'report_{timestamp}.txt')
    
    with open(report_filename, 'w') as f:
        f.write(report)
    
    print(f"Report saved as {report_filename}")
    return report

def remove_duplicates_keep_first(excel_path):
    xls = pd.ExcelFile(excel_path)
    book = load_workbook(excel_path)
    
    for sheet_name in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name)
        
        if 'Link' not in df.columns:
            possible_link_columns = [col for col in df.columns if 'link' in col.lower()]
            if possible_link_columns:
                link_column = possible_link_columns[0]
                print(f"'Link' column not found in sheet '{sheet_name}'. Using '{link_column}' instead.")
            else:
                print(f"No suitable link column found in sheet '{sheet_name}'. Skipping this sheet.")
                continue
        else:
            link_column = 'Link'
        
        df_cleaned = df.drop_duplicates(subset=link_column, keep='first')
        
        if len(df) != len(df_cleaned):
            print(f"Removed {len(df) - len(df_cleaned)} duplicate entries from sheet '{sheet_name}'")
            
            # Clear the existing sheet
            sheet = book[sheet_name]
            sheet.delete_rows(2, sheet.max_row)  # Delete all rows except the header
            
            # Write the cleaned data back to the sheet
            for r_idx, row in enumerate(df_cleaned.values, 2):
                for c_idx, value in enumerate(row, 1):
                    sheet.cell(row=r_idx, column=c_idx, value=value)
        else:
            print(f"No duplicates found in sheet '{sheet_name}'")

    book.save(excel_path)
    print(f"Saved changes to {excel_path}")

def get_region_choice():
    print("\nAvailable regions to search:")
    print("1. All Regions")
    print("2. ASIA")
    print("3. CANADA")
    print("4. EUROPE")
    print("5. OCEANIA")
    print("6. USA")
    print("7. Custom Selection")
    
    while True:
        try:
            choice = int(input("\nEnter your choice (1-7): "))
            if 1 <= choice <= 7:
                if choice == 1:
                    return None  # Return None to indicate all regions
                elif choice == 7:
                    # Custom selection
                    regions = []
                    print("\nEnter region numbers (2-6) separated by spaces:")
                    print("Example: '2 4 6' for ASIA, EUROPE, and USA")
                    selections = input("> ").split()
                    for sel in selections:
                        sel = int(sel)
                        if 2 <= sel <= 6:
                            region_map = {2: "ASIA", 3: "CANADA", 4: "EUROPE", 5: "OCEANIA", 6: "USA"}
                            regions.append(region_map[sel])
                    return regions if regions else None
                else:
                    region_map = {2: "ASIA", 3: "CANADA", 4: "EUROPE", 5: "OCEANIA", 6: "USA"}
                    return [region_map[choice]]
            else:
                print("Invalid choice. Please enter a number between 1 and 7.")
        except ValueError:
            print("Invalid input. Please enter a number.")

def load_proxies(proxy_file='proxies.txt'):
    """Load proxies from file and return as a list"""
    proxy_path = os.path.join(SCRIPT_DIR, proxy_file)
    try:
        with open(proxy_path, 'r') as f:
            proxies = [line.strip() for line in f.readlines() if line.strip()]
        print(f"Loaded {len(proxies)} proxies from {proxy_path}")
        return proxies
    except FileNotFoundError:
        print(f"Warning: {proxy_path} not found!")
        return []

def get_next_proxy(proxy_list):
    """Get next proxy from list and remove it from the file"""
    if not proxy_list:
        return None
    
    proxy = proxy_list.pop(0)  # Get and remove first proxy
    
    # Update proxies.txt with remaining proxies
    proxy_path = os.path.join(SCRIPT_DIR, 'proxies.txt')
    with open(proxy_path, 'w') as f:
        f.write('\n'.join(proxy_list))
    
    # Convert proxy string to dictionary format
    return {
        'http': f'socks5://{proxy}',
        'https': f'socks5://{proxy}'
    }

def verify_credentials(reddit_instance):
    """Verify if the Reddit instance is properly authenticated"""
    try:
        # Try to access user's own info to verify authentication
        user = reddit_instance.user.me()
        if user is None:
            return False, "Unable to get user information"
        return True, None
    except prawcore.exceptions.OAuthException:
        return False, "OAuth authentication failed"
    except prawcore.exceptions.ResponseException:
        return False, "Invalid credentials or rate limited"
    except Exception as e:
        return False, f"Authentication error: {str(e)}"

def main():
    start_time = time.time()
    all_subreddits = read_subreddits()
    
    # Load proxies at start
    proxy_list = load_proxies()
    if not proxy_list:
        print("No proxies available. Please add proxies to proxies.txt")
        return
    
    # Get user's region choice
    selected_regions = get_region_choice()
    
    # Verify all credentials before starting
    print("\nVerifying Reddit credentials...")
    valid_credentials = []
    for i, creds in enumerate(REDDIT_CREDENTIALS_LIST, 1):
        try:
            creds = creds.copy()
            if 'proxy' in creds:
                del creds['proxy']
            
            reddit = praw.Reddit(**creds)
            is_valid, error = verify_credentials(reddit)
            if is_valid:
                valid_credentials.append(creds)
                print(f"✓ Credential set {i} ({creds['username']}) verified successfully")
            else:
                print(f"✗ Credential set {i} ({creds['username']}) failed: {error}")
        except Exception as e:
            print(f"✗ Credential set {i} ({creds['username']}) failed: {str(e)}")
    
    if not valid_credentials:
        print("No valid credentials found. Please check your Reddit credentials and try again.")
        return
    
    print(f"\nProceeding with {len(valid_credentials)} valid credential sets")
    
    # Filter subreddits based on selection
    if selected_regions:
        subreddits = {region: all_subreddits[region] for region in selected_regions}
        print(f"\nSelected regions: {', '.join(selected_regions)}")
    else:
        subreddits = all_subreddits
        print("\nSearching all regions")
    
    link_dict = {region: {format: set() for format in LINK_FORMATS} for region in subreddits.keys()}
    skipped_subreddits = []
    
    current_credential_index = 0
    requests_made = 0
    current_proxy = None

    for region, region_subreddits in subreddits.items():
        print(f"\nProcessing region: {region}")
        excel_path = os.path.join(SCRIPT_DIR, f'links{region}.xlsx')
        
        try:
            book = load_workbook(excel_path)
            print(f"Loaded existing file: {excel_path}")
        except FileNotFoundError:
            book = Workbook()
            book.remove(book.active)
            print(f"Created new workbook for: {excel_path}")

        for format in LINK_FORMATS:
            if format not in book.sheetnames:
                book.create_sheet(format)
                book[format].append(['Link', 'Subreddit', 'Date Added'])

        total_subreddits = len(region_subreddits)
        for index, subreddit in enumerate(region_subreddits, start=1):
            # Check if we need to rotate credentials and get new proxy
            if requests_made >= REQUESTS_PER_CREDENTIAL:
                current_credential_index = (current_credential_index + 1) % len(valid_credentials)
                requests_made = 0
                current_proxy = get_next_proxy(proxy_list)  # Get new proxy only when rotating credentials
                if not current_proxy:
                    print("No more proxies available! Please add more proxies to proxies.txt")
                    return
                print(f"\nRotating to credential set {current_credential_index + 1} ({valid_credentials[current_credential_index]['username']})")
                time.sleep(2)
            
            # Initialize Reddit instance with current credentials and proxy
            try:
                if current_proxy is None:  # First run
                    current_proxy = get_next_proxy(proxy_list)
                    if not current_proxy:
                        print("No more proxies available! Please add more proxies to proxies.txt")
                        return

                current_creds = valid_credentials[current_credential_index].copy()
                current_creds['proxy'] = current_proxy
                reddit = praw.Reddit(**current_creds)
            except Exception as e:
                print(f"Error initializing Reddit instance: {str(e)}")
                continue
            
            print(f"Searching subreddit {index}/{total_subreddits} in {region}: {subreddit}")
            links, error_message = search_subreddit(reddit, subreddit, SEARCH_KEYWORDS, START_DATE, END_DATE)
            requests_made += 1

            if error_message:
                print(f"Skipped {subreddit}. Reason: {error_message}")
                skipped_subreddits.append((subreddit, error_message))
            else:
                for link in links:
                    for format in LINK_FORMATS:
                        if format in link:
                            update_excel_sheet(book[format], [link], subreddit)
                            link_dict[region][format].add(link)
                            break

            time.sleep(2)  # Base delay between requests

        book.save(excel_path)
        print(f"Saved all links for {region} to the Excel file: {excel_path}")
        remove_duplicates_keep_first(excel_path)

    end_time = time.time()
    report = generate_report(link_dict, skipped_subreddits, start_time, end_time, START_DATE, END_DATE)
    print("\nSummary Report:")
    print(report)

if __name__ == "__main__":
    main()